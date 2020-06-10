# Author: Chris Paroly
# Title: Netowrk Engineer @ Network Services
# Contact: cparoly@northwell.edu
# Reporting to : Mgr. Thomas Hynes

import openpyxl
from netmiko import ConnectHandler
from netmiko import NetMikoTimeoutException
from netmiko import NetMikoAuthenticationException

devices = []
count = 1
book = openpyxl.load_workbook('devices.xlsx')
sheet = book.active

for column in sheet['A']:
	devices.append(column.value)


ips = iter(devices)
next(ips)


config_commands = ['username nwhl@nloc password 0 testing']

for ip in ips:
	count += 1
	print ("test" + str(count))
	device = ''.join(ip)
	cisco = {
			'device_type': 'cisco_ios',
			'host': device,
			'username': 'cisco',
			'password': 'cisco',
		}
	print('Connecting to device ' + device.strip() + ' ... Please Wait \n')
	file = open(str(device.strip('\n'))+'.txt', 'w')
	try:
		net_connect = ConnectHandler(**cisco)
	except NetMikoTimeoutException:
		print("Device unreachable")
		cell = sheet.cell(row=count, column=2)
		cell.value = "Unreachable"
		continue

	except NetMikoAuthenticationException:
		print("Authentication Error")
		cell = sheet.cell(row=count, column=2)
		cell.value = "Authentication Error"
		continue


	print(net_connect.find_prompt())

	output = net_connect.send_config_set(config_commands)

	print(output)
	net_connect.send_command("write mem")
	net_connect.send_command("\n")

	cell = sheet.cell(row=count, column=2)
	cell.value = "success"

	file.write(output)
	file.close()

book.save('devices.xlsx')
