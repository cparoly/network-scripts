from netmiko import ConnectHandler
from netmiko import NetMikoAuthenticationException
import openpyxl
import re
from mac_vendor_lookup import MacLookup

vendor_lookup = MacLookup()
book = openpyxl.load_workbook("devices.xlsx")
sheet = book.active
sheet["F1"] = "Status"
sheet["G1"] = "MAC"
sheet["H1"] = "Vendor"
hosts = []
count = 2
mac_regex = re.compile(r'[0-9a-f]{4}\.[0-9a-f]{4}\.[0-9a-f]{4}')
switch = {
	'device_type': 'cisco_ios',
	'host': "172.16.1.3",
	'username': "test",
	'password': "test"
}
try:
	dist_connect = ConnectHandler(**switch)
except NetMikoAuthenticationException:
	print("username/password is incorrect")



for column in sheet["C"]:
	hosts.append(column.value)

hosts = iter(hosts)
next(hosts)

for ip in hosts:
	arp = dist_connect.send_command("show ip arp " + str(ip))
	mac = re.search(mac_regex, arp)

	if mac is not None:
		sheet.cell(row=count, column = 6, value="Online")
		mac = mac.group()
		sheet.cell(row=count, column = 7, value= mac)
		vendor = vendor_lookup.lookup(mac)
		sheet.cell(row=count, column= 8, value = vendor)
		count += 1

	else:
		sheet.cell(row=count, column = 6, value="Inactive")
		count += 1

book.save("devices.xlsx")
	

