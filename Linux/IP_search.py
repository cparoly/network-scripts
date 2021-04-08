

import openpyxl
from openpyxl import Workbook
import socket
from netmiko import ConnectHandler
from netmiko import NetMikoTimeoutException
from netmiko import NetMikoAuthenticationException
import re
import ipaddress
import getpass
import os
# get the user's home folder
homedir = os.path.expanduser('~')

mac_regex = re.compile(r'[0-9a-f]{4}\.[0-9a-f]{4}\.[0-9a-f]{4}')
int_regex = re.compile(r'Fa{1}\S*\d/\S*\d{1,2}|Gi{1}\S*\d/\S*\d|Eth{1}\d/\S*\d{1,2}|Te{1}\S*\d/\S*\d')
int_po_regex = re.compile(r'Po{1}\d*')
int_regexes = [int_regex, int_po_regex]
description_regex = re.compile(r'Description: (.*)', re.MULTILINE)
access_vlan_regex = re.compile(r'switchport access vlan (\d*)', re.MULTILINE)

wb = Workbook()
ws = wb.active

ws['A1'] = 'IPs'
ws['B1'] = 'Switch'
ws['C1'] = 'Interface'
ws['D1'] = 'Duplex'
ws['E1'] = 'Speed'
ws['F1'] = 'CRC'
ws['G1'] = 'Input Errors'
ws['H1'] = 'Output Drops'

user = input("Tacacs username: ")
pwd = getpass.getpass("Tacacs password")

while True:
    try:
        print("Make sure the file is in the Documents folder")
        file = input("What is the name of the file (include the .xlsx): ")
        book = openpyxl.load_workbook(f'{homedir}/Documents/{file}')

        sheet = book.active
        break
    except ValueError:
        print("Wrong file path")
        continue

col = input("What column number are the IPs? ")
try:
    distribution = input("What is the distribution switch that will have the ARP table for these? ")
    distribution = ipaddress.ip_address(distribution)
except ValueError:
    distribution = socket.gethostbyname(distribution)

    cisco = {
        'device_type': 'cisco_ios',
        'host': distribution,
        'username': user,
        'password': pwd
    }
    try:
        dist_connect = ConnectHandler(**cisco)
    except NetMikoAuthenticationException:
        print("Username/password incorrect")

list_ips = []

count = 1
# get list of IPs from the file
for column in sheet["A"]:
    list_ips.append(column.value)

list_ips = iter(list_ips)
next(list_ips)
ips = list(filter(None, list_ips))


def ConnectToDevice(host):
    # Connects to the devices
    print("Connecting to " + host)
    switch = {
        'device_type': 'cisco_ios',
        'host': host,
        'username': user,
        'password': pwd
    }
    try:
        net_connect = ConnectHandler(**switch)
    except NetMikoAuthenticationException:
        raise

    except NetMikoTimeoutException:
        raise

    return net_connect


def GetMac(ip):
    # get the mac address associated with the IP from the arp table
    print('*' * 50)
    print('Getting the MAC ADDRESS')
    try:
        arp = dist_connect.send_command("show ip arp " + ip, delay_factor=.2)
        print(arp)
        mac = re.search(mac_regex, arp)
        if mac is None:
            print('No arp entry found')
            ws.cell(row=count, column=3, value='NONE')
            return None
        else:
            mac = mac.group()
            return mac
    except:
        pass


def GetNextSwitch(mac):
    # Gets the switch the mac is connected to
    print('*' * 50)
    print('Getting next switch')
    int = dist_connect.send_command("show mac address-table address " + mac, delay_factor=.1)
    # Checks to see if it is a port-channel or regular interface
    po = re.search(int_po_regex, int)
    if po is None:
        # looks for the interface since it wasn't a port-channel
        po = re.search(int_regex, int)

        if po is None:
            return None
        else:
            po = po.group()
    else:
        po = po.group()

    print(po)
    # find the next switch by reading the description of the Port-Channel or Interface
    interface = dist_connect.send_command("show run int " + str(po))
    print(interface)
    for description in interface.splitlines():
        if 'description' in description.lower():
            next_switch = description.split(' ')

    next_switch = [switch for switch in next_switch if 'ns-s' in switch or 'nw-s' in switch]
    next_switch = str(next_switch)[1:-1]
    next_switch = next_switch.strip("'")
    print(next_switch)

    return next_switch


def FindAttachedInterface(switch, mac):
    # looks for the mac's interface to supply interface information
    try:
        net_connect = ConnectToDevice(switch)
    except NetMikoAuthenticationException:
        return "Failed to Connect"
    except NetMikoTimeoutException:
        return "Failed to Connect"
    except ValueError:
        print(ValueError)
        return "Failed to Connect"
    interface = net_connect.send_command("show mac address-table address " + mac)
    print(interface)
    int = re.search(int_regex, interface)
    # if the mac isn't found
    if int is None:
        int = re.search(r'Po{1}\d*$', interface)
        if int is None:
            interface = 'unknown'
            duplex = 'unknown'
            speed = 'unknown'
            CRC = 'unknown'
            input_error = 'unknown'
            output_drops = 'unknown'
            return interface, duplex, speed, CRC, input_error, output_drops
        else:
            int = int.group()
    else:
        int = int.group()

    interfaces = net_connect.send_command("show int " + str(int), use_genie=True)
    # supply us with all needed information from the interface
    for interface in interfaces:
        bandwidth = interfaces[interface]['bandwidth']
        duplex = interfaces[interface]['duplex_mode']
        speed = interfaces[interface]['port_speed']
        CRC = interfaces[interface]['counters']['in_crc_errors']
        input_error = interfaces[interface]['counters']['in_errors']
        output_drops = interfaces[interface]['queues']['total_output_drop']

    net_connect.disconnect()
    return interface, duplex, speed, CRC, input_error, output_drops


print("Connecting to " + str(distribution))

for ip in ips:
    # runs through the list of IPs from the excel file and starts the process of getting the mac address
    # and tracking down the interface it is connected to.
    print(ip)
    count += 1
    mac = GetMac(ip)
    if mac is None:
        print('No arp entry found')
        ws.cell(row=count, column=3, value='NONE')
        continue
    switch = GetNextSwitch(mac)
    if switch is None:
        print('No interface found---GetNetworkSwitch')
        ws.cell(row=count, column=3, value='NONE')
        continue
    print('The host is connected to ' + switch)
    if FindAttachedInterface(switch, mac) == 'Failed to Connect':
        continue
    interface, duplex, speed, CRC, input_error, output_drops = FindAttachedInterface(switch, mac)
    if interface is None:
        print('No Interface found----FindAttachedInterface')
        ws.cell(row=count, column=3, value='NONE')
        continue
    ws.cell(row=count, column=1, value=ip)
    ws.cell(row=count, column=3, value=interface)
    ws.cell(row=count, column=4, value=duplex)
    ws.cell(row=count, column=5, value=speed)
    ws.cell(row=count, column=2, value=switch)
    ws.cell(row=count, column=6, value=CRC)
    ws.cell(row=count, column=7, value=input_error)
    ws.cell(row=count, column=8, value=output_drops)

input("Press enter to exit")
wb.save(filename=f'{homedir}/Documents/host-interfaces.xlsx')
dist_connect.disconnect()
book.close()
