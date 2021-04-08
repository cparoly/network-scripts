# Created by Chris Paroly


from netmiko import ConnectHandler
from netmiko import NetMikoTimeoutException
from netmiko import NetMikoAuthenticationException
from openpyxl import Workbook
import socket
import getpass
import os

homedir = os.path.expanduser('~')

# Opens Excel file



def get_mac_address():
    switch_mac = []
    switch_port = []
    switch_mac_table = {}
    print("******" * 50)
    print(switch_mac_table)
    # Gets the mac address table from the switch and gives a dictionary with Mac address and the Port
    print("getting MAC table from " + mapping)
    print(switch_connect.find_prompt())
    mac_address = switch_connect.send_command('show mac address-table', use_textfsm=True)
    for switch_macs in mac_address:
        if switch_macs['destination_port'].lower().startswith(('v', 'p', 'c')):
            continue
        else:
            switch_mac.append(switch_macs['destination_address'])
            switch_port.append(switch_macs['destination_port'])
    switch_mac_table = dict(zip(switch_mac, switch_port))
    print(switch_mac_table)
    return switch_mac_table


def collect_information(mac_address_table):
    # get int status to get the interfaces information
    print(switch_connect.find_prompt())
    status = switch_connect.send_command('show int status', use_textfsm=True)
    count = 2
    for interface in status:
        try:
            port = interface['port']
            desc = interface['name']
            vlan = interface['vlan']
            status = interface['status']
        except TypeError as t:
            print(t)
            continue
        # Adds that information to the excel spreadsheet
        ws.cell(row=count, column=1, value=str(port))
        ws.cell(row=count, column=2, value=str(status))
        ws.cell(row=count, column=3, value=str(desc))
        ws.cell(row=count, column=4, value=str(vlan))

        for address, mac_port in mac_address_table.items():
            if "GigabitEthernet" in mac_port:
                mac_port = mac_port.replace("GigabitEthernet", "Gi")
            elif "FastEthernet" in mac_port:
                mac_port = mac_port.replace("FastEthernet", "Fa")
            else:
                pass
            if mac_port == port:
                # finds the mac address per port
                print(mac_port, port)
                ws.cell(row=count, column=5, value=address)
                for mac_address, ip_address in arp_table.items():
                    # find the ip address per mac address
                    if mac_address == address:
                        ws.cell(row=count, column=6, value=ip_address)
                        try:
                            # gets the hostname for each IP address and adds it to the excel, or value is NONE
                            hostname = socket.gethostbyaddr(ip_address)
                            ws.cell(row=count, column=7, value=str(hostname[0]))
                        except Exception:
                            ws.cell(row=count, column=7, value="NONE")

        # counts for each row
        count += 1

    # saves the excel to the Documents folder for each switch
    wb.save(f"{homedir}/Documents/{mapping}-port-mapping.xlsx")
    wb.close()


switches = []

user = input("Tacacs username: ")
pwd = getpass.getpass("Tacacs password")

dist_switch = input("What is the distribution switch to get the ARP table from?: ")
while True:
    switch = input("What switch are you scanning?: ")
    switches.append(switch)
    again = input("Is there another switch you are scanning? (Y/N) ")
    if again.lower() == 'y':
        continue
    else:
        break

cisco = {
    'device_type': 'cisco_ios',
    'host': dist_switch,
    'username': user,
    'password': pwd
}

ips = []
macs = []

try:
    net_connect = ConnectHandler(**cisco)
except NetMikoAuthenticationException:
    print('Wrong Username or Password')
except NetMikoTimeoutException:
    print("Connection to Switch failed")

arp = net_connect.send_command("show ip arp ", use_genie=True)
# Does a show ip arp command and produces all mac address and IP address
for key in arp:
    for value in arp[key]:
        try:
            for neighbor in arp[key][value]['ipv4']['neighbors']:
                mac = arp[key][value]['ipv4']['neighbors'][neighbor]['link_layer_address']
                macs.append(mac)

                ip = arp[key][value]['ipv4']['neighbors'][neighbor]['ip']
                ips.append(ip)
        except KeyError:
            ip = arp[key][value]['ip_address']
            ips.append(ip)

            mac = arp[key][value]['mac_address']
            macs.append(mac)

arp_table = dict(zip(macs, ips))

for mapping in switches:
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Interface'
    ws['B1'] = 'Status'
    ws['C1'] = 'Description'
    ws['D1'] = 'Vlan'
    ws['E1'] = 'MAC'
    ws['F1'] = 'IP Address'
    ws['G1'] = 'Hostname'

    # loops over each switch in the list given
    switch = {
        'device_type': 'cisco_ios',
        'host': mapping,
        'username': user,
        'password': pwd
    }

    try:
        print("connecting to " + mapping)
        switch_connect = ConnectHandler(**switch)
    except NetMikoAuthenticationException:
        print('Wrong Username or Password')
        break
    except NetMikoTimeoutException:
        print("Connection to Switch failed")
        continue
    print("sending to get_mac_address")
    mac_table = get_mac_address()
    print("sending to collect_information")
    collect_information(mac_table)

print("Check the Documents folder for the excel spreadsheets")
