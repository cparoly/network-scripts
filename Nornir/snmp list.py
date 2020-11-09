from nornir import InitNornir
from nornir.plugins.functions.text import print_result
from nornir.core.filter import F
import openpyxl
from openpyxl import Workbook

nr = InitNornir(core={"num_workers": 100},
                inventory={
                    "plugin": "nornir.plugins.inventory.simple.SimpleInventory",
                    "options": {
                        "host_file": "inventory/hosts.yaml",
                        "group_file": "inventory/groups.yaml",
                        "default_file": "inventory/defaults.yaml"
                    }
                }
                )
wb = Workbook()

file = input("Enter the name for the excel spreadsheet: ")
file = file + ".xlsx"
ws = wb.active

ws["A1"] = "Hosts"
row = 2

hospitals = ["Forest Hills Hospital", "Glen Cove Hospital", "Huntington Hospital", "LHH", "LIJ", "Mather Hospital",
             "MEETH", "MIT", "NSUH", "NWH", "PBMC", "Phelps Hospital", "Plainview Hospital", "SIUH",
             "South Oaks Hospital", "Southshore Hospital", "Syosset Hospital", "Valley Stream Hospital", "Westbury", ]


def print_hospitals():
    print('1. Forest Hills Hospital')
    print('2. Glen Cove Hospital')
    print('3. Huntington Hospital')
    print('4. "LHH')
    print('5. LIJ')
    print('6. Mather Hospital')
    print('7. MEETH')
    print('8. MIT')
    print('9. NSUH')
    print('10. NWH')
    print('11. PBMC')
    print('12. Phelps Hospital')
    print('13. Plainview Hospital')
    print('14. SIUH')
    print('15. South Oaks Hospital')
    print('16. Southshore Hospital')
    print('17. Syosset Hospital')
    print('18. Valley Stream Hospital')
    print('19. Westbury')


print(print_hospitals())

while True:
    count = input('How many sites are you updating? (1-3): ')
    count = int(count)
    if count == 1:
        selection = input("Please choose which hospital to update from the list above by their number: ")
        selection = hospitals[int(selection) - 1]
        print('The following will be updated ' + selection)
        location = nr.filter(F(site=selection))
        break
    elif count == 2:
        selection1 = input("Please choose the first hospital to update from the list above by their number: ")
        selection1 = hospitals[int(selection1) - 1]
        selection2 = input("Select the second hospital: ")
        selection2 = hospitals[int(selection2) - 1]
        print('The following will be updated ' + selection1 + ' and ' + selection2)
        location = nr.filter(F(site=selection1) | F(site=selection2))
        break
    elif count == 3:
        selection1 = input("Please choose the first hospital to update from the list above by their number: ")
        selection1 = hospitals[int(selection1) - 1]
        selection2 = input("Select the second hospital: ")
        selection2 = hospitals[int(selection2) - 1]
        selection3 = input("Select the third hospital: ")
        selection3 = hospitals[int(selection3) - 1]
        print('The following will be updated ' + selection1 + ', ' + selection2 + " and " + selection3)
        location = nr.filter(F(site=selection1) | F(site=selection2) | F(site=selection3))
        break
    else:
        print("Invalid input")
        continue
print('--' * 40)
devices = location.filter(F(type="Cisco IOS - SSH Capable") | F(type="Cisco NX OS") | F(type="Telnet")).inventory.hosts.keys()

for device in devices:
    print(device)
    ws.cell(row=row, column=1, value=device)
    row += 1

wb.save(file)
