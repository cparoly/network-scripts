import openpyxl

hosts = open(r'C:\Users\CParoly\PycharmProjects\work\Nornir\inventory\lij.yaml', 'w')
hosts.write('---')
wb = openpyxl.load_workbook("LIJ_Swith_Router_List.xlsx")
ws = wb.active
count = 2
for row in ws.rows:

	device = ws.cell(row=count, column=1).value
	ip = ws.cell(row=count, column=2).value
	platform = ws.cell(row=count, column=3).value
	location = ws.cell(row=count, column=4).value
	ios = ws.cell(row=count, column =5).value
	building = ws.cell(row=count, column=6).value

	hosts.write("\n{0}:\n  hostname: {1}\n  groups:\n    - {2}\n  data:\n"
				"    site: {3}\n    type: {4}\n    building: {5}"
				.format(device, ip, platform, location, ios, building))
	count += 1

hosts.close()


