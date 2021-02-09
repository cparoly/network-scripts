from netmiko import ConnectHandler
from nornir import InitNornir
from nornir.plugins.tasks import commands
from nornir.plugins.functions.text import print_result
from nornir.plugins.tasks.networking import netmiko_send_command
from nornir.plugins.tasks.networking import netmiko_send_config
from nornir.plugins.tasks.networking import napalm_configure
from nornir.core.filter import F
import openpyxl


book = openpyxl.load_workbook("network devices.xlsx")
sheet = book.active
sheet["F15"] = "System"
row = 123
nr = InitNornir(core={"num_workers": 100},
                inventory={
                    "plugin": "nornir.plugins.inventory.simple.SimpleInventory",
                    "options": {
                        "host_file": "inventory/hosts.yaml",
                        "group_file": "inventory/groups.yaml",
                        "default_file": "inventory/defaults.yaml"
                    }
                }
                )

testing = nr.filter(F(groups__contains='Router')).filter(site='MIT')
print('Starting')
show = testing.run(task=netmiko_send_command,
                   command_string='show version | i Software')
print_result(show)

for host in show:
    #f = open(f"{host}.txt", "w")

    IOS = show[host][0].result
    print(host)
    print(IOS)
    if "Nexus" in IOS:
        sheet.cell(row=row, column=6, value="Nexus")
        row += 1
    elif "Internetwork" in IOS:
        sheet.cell(row=row, column=6, value="Internetwork")
        row += 1
    elif "XE " in IOS:
        sheet.cell(row=row, column=6, value="IOS-XE")
        row += 1
        print('IOS-XE')
        print(row)
        print("---" * 30)
    elif "Adaptive Security Appliance" in IOS:
        sheet.cell(row=row, column=6, value="ASA")
        row += 1
        print("---" * 30)
    elif "IOS " in IOS:
        sheet.cell(row=row, column=6, value="IOS")
        row += 1
        print('IOS')
        print(row)
        print("---"*30)
    else:
        row += 1
        print(row)
        print("---" * 30)
        continue

book.save("network devices.xlsx")
