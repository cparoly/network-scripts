from nornir import InitNornir
from nornir_netmiko.tasks import netmiko_send_command, netmiko_send_config
from nornir_jinja2.plugins.tasks import template_file
from nornir_utils.plugins.functions import print_result
from nornir.core.filter import F
import openpyxl
import getpass
from tqdm import tqdm

wb = openpyxl.load_workbook("all network devices in Spectrum.xlsx")
ws = wb.active

user = input("Enter your tacacs username: ")
password = getpass.getpass(prompt="Enter your tacacs password (will not show): ")

hospitals = ["Forest Hills Hospital", "Glen Cove Hospital", "Huntington Hospital", "LHH", "LIJ", "Mather Hospital",
             "MEETH", "MIT", "NSUH", "NWH", "PBMC", "Phelps Hospital", "Plainview Hospital", "SIUH",
             "South Oaks Hospital", "Southshore Hospital", "Syosset Hospital", "Valley Stream Hospital", "Westbury",
             "Remote"]

nr = InitNornir(config_file='config.yaml')

nr.inventory.defaults.username = user
nr.inventory.defaults.password = password


def print_hospitals():
    print('1. Forest Hills Hospital')
    print('2. Glen Cove Hospital')
    print('3. Huntington Hospital')
    print('4. LHH')
    print('5. LIJ')
    print('6. Mather Hospital')
    print('7. MEETH')
    print('8. MIT')
    print('9. NSUH')
    print('10. NWH')
    print('11. PBMC')
    print('12. Phelps Hospital')
    print('13. Plainview Hospital')
    print('14. SIUH')
    print('15. South Oaks Hospital')
    print('16. Southshore Hospital')
    print('17. Syosset Hospital')
    print('18. Valley Stream Hospital')
    print('19. Westbury')
    print('20. Remote sites')


def results_file(completed):
    # creates txt file with configuration output
    for host in completed:
        f = open(f"{host}.txt", "w")
        f.write(str(completed[host].result[0]))
    else:
        print("none")


def to_excel(device):
    for host in device:
        for cell in ws['A']:
            if cell.value is not None:
                if host in cell.value:
                    row = cell.row

        if device[host].failed is True:
            ws.cell(row=row, column=10, value="Failed")
        else:
            ws.cell(row=row, column=10, value="Success")


def replace_users(task):
    # Finds username doing a "show run | i username" against the switch
    usernames = task.run(task=netmiko_send_command, command_string='show run | i username')
    usernames = usernames.result
    # splits the output into a list per line and removes the aaa entry it will find
    usernames = usernames.split('\n')
    try:
        usernames.remove('aaa authentication username-prompt Username:')
    except ValueError:
        pass
    filter(None, usernames)
    task.host['users'] = usernames
    # adds the list of usernames to the template
    template = task.run(task=template_file, template="username.j2", path=f"./templates/{task.host['building']}/{task.host.platform}")
    user_template = template.result
    return user_template


def run_template(task, progress):
    #config = replace_users(task)
    #result = task.run(task=netmiko_send_config, config_commands=config, delay_factor=2, max_loops=400, cmd_verify=False)
    result = task.run(task=netmiko_send_command, command_string='copy running-config startup-config')
    progress.update()
    return result


print(print_hospitals())
print("----" * 40)
while True:
    count = input('How many sites are you updating? (1-3): ')
    count = int(count)
    if count == 1:
        selection = input("Please choose which location to update from the list above by their number: ")
        selection = hospitals[int(selection) - 1]
        print('The following will be updated ' + selection)
        yesno = input("Press enter if that is correct entry. (Y/N)")
        if yesno.lower() == "n":
            continue
        elif selection == "Remote sites":
            location = nr.filter(F(building='Remote'))
            break
        else:
            location = nr.filter(F(site=selection))
            break
    elif count == 2:
        selection1 = input("Please choose the first location to update from the list above by their number: ")
        selection1 = hospitals[int(selection1) - 1]
        selection2 = input("Select the second hospital: ")
        selection2 = hospitals[int(selection2) - 1]
        print('The following will be updated ' + selection1 + ' and ' + selection2)
        yesno = input("Press enter if that is correct entry. (Y/N)")
        if yesno.lower() == "n":
            continue
        else:
            location = nr.filter(F(site=selection1) | F(site=selection2))
            break
    elif count == 3:
        selection1 = input("Please choose the first location to update from the list above by their number: ")
        selection1 = hospitals[int(selection1) - 1]
        selection2 = input("Select the second hospital: ")
        selection2 = hospitals[int(selection2) - 1]
        selection3 = input("Select the third hospital: ")
        selection3 = hospitals[int(selection3) - 1]
        print('The following will be updated ' + selection1 + ', ' + selection2 + " and " + selection3)
        yesno = input("Press enter if that is correct entry. (Y/N)")
        if yesno.lower() == "n":
            continue
        else:
            location = nr.filter(F(site=selection1) | F(site=selection2) | F(site=selection3))
            break
    else:
        print("Invalid input")
        continue

# test = nr.filter(F(site='3hq'))
#devices = location.filter(F(type="Cisco IOS - SSH Capable") | F(type="Cisco NX OS") | F(type="Telnet") | F(type="DMVPN"))
devices = location.filter(type='Cisco NX OS')
with tqdm(total=len(devices.inventory.hosts), desc="Updating ") as progress_bar:
    complete = devices.run(task=run_template, progress=progress_bar)
print_result(complete)
# to_excel(complete)
# wb.save('all network devices in Spectrum.xlsx')
# results_file(complete)
