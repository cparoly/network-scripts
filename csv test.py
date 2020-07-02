import openpyxl
import socket

book = openpyxl.load_workbook('devices.xlsx')

sheet = book.active

devices = []
count = 1
for column in sheet['A']:

	devices.append(column.value)
	print(devices)
	# cell = sheet.cell(row= count, column = 2)
	# cell.value = ip
	print("test" + str(count))
	count += 1

ips = iter(devices)
next(ips)


count = 2
for ip in ips:
	print (ip)
	hostname = socket.gethostbyname(ip)
	print(hostname)

	cell = sheet.cell(row= count, column = 2)
	cell.value = hostname
	count += 1
book.save('devices.xlsx')


